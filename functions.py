import requests
from bs4 import BeautifulSoup
import openpyxl
import re
from datetime import datetime, date
from openpyxl.styles import colors
from openpyxl.styles import Font, PatternFill
import math


def get_prices():

    # Examples of URLs (⚠️ Kabum may change the urls if they were in sale ⚠️)
    # when you change the urls, remember to change the name of the row in the sheets.xlsx

    urls = ['https://www.kabum.com.br/produto/102249/processador-amd-ryzen-5-3400g-cache-6mb-3-7ghz-4-2ghz-max-turbo-am4-yd3400c5fhbox',
            'https://www.kabum.com.br/produto/95677/placa-mae-asus-ex-a320m-gaming-amd-am4-matx-ddr4',
            'https://www.kabum.com.br/cgi-local/site/produtos/descricao_ofertas.cgi?codigo=84108',
            'https://www.kabum.com.br/cgi-local/site/produtos/descricao_ofertas.cgi?codigo=103547',
            'https://www.kabum.com.br/produto/102248/processador-amd-ryzen-3-3200g-cache-4mb-3-6ghz-4ghz-max-turbo-am4-yd3200c5fhbox',
            'https://www.kabum.com.br/produto/99927/gabinete-gamer-nox-hummer-tgm-rgb-rainbow-4-coolers-lateral-e-frontal-em-vidro-nxhummertgm',
            'https://www.kabum.com.br/produto/100626/teclado-gamer-dazz-rapid-fire-revolution-rainbow-abnt2-625203',
            'https://www.kabum.com.br/produto/111939/monitor-lg-led-19-5-hdmi-vga-2ms-ajuste-de-inclinacao-20mk400h-b',
            'https://www.kabum.com.br/produto/91021/fonte-corsair-450w-80-plus-bronze-cx450-cp-9020120-br',
            ]

    prices = []

    for url in urls:
        try:
            response = requests.get(url)
            html = BeautifulSoup(response.text, 'html.parser')
            if html.select_one('.preco_desconto_avista-cm'):
                price = re.sub('[R$àvista]', '', html.select_one('.preco_desconto_avista-cm').text).strip()
                price = re.sub('\.', '', price)
                price = math.ceil(float(re.sub(',', '.', price)))
            elif html.select_one('.preco_desconto-cm'):
                price = re.sub('[R$àvista]', '', html.select_one('.preco_desconto-cm strong').text).strip()
                price = re.sub('\.', '', price)
                price = math.ceil(float(re.sub(',', '.', price)))
            else:
                price = re.sub('[R$àvista]', '', html.select_one('.preco_desconto strong').text).strip()
                price = re.sub('\.', '', price)
                price = math.ceil(float(re.sub(',', '.', price)))
        except AttributeError:
            price = "INDISPONÍVEL"
        finally:
            prices.append(price)

    return prices


def start_day_file():
    day = 1 
    with open('day.txt', 'w') as file:
        file.write(str(day + 1) + '\n')
        file.write(date.today().strftime("%d/%m/%Y"))
    return 1


def add_day():
    day = int(get_column_from_day())
    with open('day.txt', 'w') as file:
        file.write(str(day+1)+'\n')
        file.write(date.today().strftime("%d/%m/%Y"))


def get_column_from_day():
    with open('day.txt', 'r+') as file:
        return int(file.readline())


def get_day_month():
    return f'{datetime.today().day}/{datetime.today().month}'


def validate_today_date():
    with open('day.txt', 'r+') as file:
        try:
            if not file.readlines:
                return start_day_file()
            else:
                date_from_file = file.readlines()[1]
                date_from_file = datetime(int(date_from_file[6:]), 
                                          int(date_from_file[3:5]), 
                                          int(date_from_file[:2])).date()

                return False if not date_from_file < datetime.now().date() else True
        except IndexError:
            return start_day_file()


def save_into_sheet():
    font_b = Font(name="Calibri", size=14, color=colors.BLACK)
    font_w = Font(name="Calibri", size=14, color=colors.WHITE)
    font_w_bold = Font(name="Calibri", size=14, color=colors.WHITE, bold=True)
    col = get_column_from_day()
    prices = get_prices()
    sheets = openpyxl.load_workbook('sheets.xlsx')
    sheet = sheets.active
    sheet.cell(row=2, column=col).value = get_day_month()
    sheet.cell(row=2, column=col).font = font_w
    for i, x in enumerate(range(3,11)):
        sheet.cell(row=x, column=col).value = prices[i]
        if i >= 0 and i < 4:
            sheet.cell(row=x, column=col).font = font_w_bold
            sheet.cell(row=x, column=col).fill = PatternFill("solid", fgColor="00FF0000")
        else:
            sheet.cell(row=x, column=col).font = font_b
    sheets.save('sheets.xlsx')