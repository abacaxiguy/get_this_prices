import requests
from bs4 import BeautifulSoup
import openpyxl
import re
from datetime import datetime
from datetime import date
from openpyxl.styles import colors
from openpyxl.styles import Font


def get_prices():

    urls = ['https://www.kabum.com.br/produto/102249/processador-amd-ryzen-5-3400g-cache-6mb-3-7ghz-4-2ghz-max-turbo-am4-yd3400c5fhbox',
            'https://www.kabum.com.br/produto/95677/placa-mae-asus-ex-a320m-gaming-amd-am4-matx-ddr4',
            'https://www.kabum.com.br/produto/84108/hd-seagate-barracuda-1tb-3-5-sata-st1000dm010?',
            'https://www.kabum.com.br/cgi-local/site/produtos/descricao_ofertas.cgi?codigo=103547',
            'https://www.kabum.com.br/produto/91021/fonte-corsair-450w-80-plus-bronze-cx450-cp-9020120-br',
            'https://www.kabum.com.br/produto/99927/gabinete-gamer-nox-hummer-tgm-rgb-rainbow-4-coolers-lateral-e-frontal-em-vidro-nxhummertgm',
            ]
    prices = []

    for url in urls:
        try:
            response = requests.get(url)
            html = BeautifulSoup(response.text, 'html.parser')
            if html.select_one('.preco_desconto_avista-cm'):
                price = re.sub('[R$àvista]','', html.select_one('.preco_desconto_avista-cm').text).strip()
            else:
                price = re.sub('[R$àvista]','', html.select_one('.preco_desconto strong').text).strip()
        except AttributeError:
            price = 'Erro'
        finally:
            prices.append(price)

    return prices


def add_day():
    day = int(get_column_from_day())
    with open('day.txt', 'w') as file:
        file.write(str(day+1)+'\n')
        file.write(date.today().strftime("%d/%m/%Y"))


def get_column_from_day():
    with open('day.txt', 'r+') as file:
        return int(file.readline())


def get_day_month():
    return f'{datetime.today().day}/{datetime.today().month}'


def validate_today_date():
    with open('day.txt', 'r+') as file:
        if not file.readlines()[1] < date.today().strftime("%d/%m/%Y"):
            return 0
        else: return 1


def save_into_sheet():
    font_b = Font(name="Calibri", size=14, color=colors.BLACK)
    font_w = Font(name="Calibri", size=14, color=colors.WHITE)
    col = get_column_from_day()
    prices = get_prices()
    sheets = openpyxl.load_workbook('sheets.xlsx')
    sheet = sheets.active
    sheet.cell(row=2, column=col).value = get_day_month()
    sheet.cell(row=2, column=col).font = font_w
    for i, x in enumerate(range(3,9)):
        sheet.cell(row=x, column=col).value = prices[i]
        sheet.cell(row=x, column=col).font = font_b
    sheets.save('sheets.xlsx')


if validate_today_date():
    save_into_sheet()
    add_day()